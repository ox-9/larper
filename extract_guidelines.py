#!/usr/bin/env python3
"""
PDF Guideline Extractor - Conservative Approach
Extracts only actual guidelines/rules, not every sentence.
"""

import pdfplumber
import re
import json
import os
import sys
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Optional, Tuple, Set
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


@dataclass
class GuidelineSection:
    """Represents a single guideline section from the PDF."""
    id: str
    category: str
    title: str
    content: str
    page_number: int
    level: int = 0
    parent_id: Optional[str] = None
    guideline_type: str = "general"  # 'requirement', 'prohibition', 'threshold', 'procedure', 'general'


class PDFGuidelineExtractor:
    """Extracts structured guidelines from PDF documents - CONSERVATIVE approach."""

    # Patterns that indicate actual guidelines/requirements
    GUIDELINE_INDICATORS = [
        # Must/Shall/Required statements
        r'\b(?:must|shall|required|need to|have to|is required)\b',
        # Prohibitions
        r'\b(?:may not|shall not|must not|cannot|prohibited|not allowed|no \w+ allowed)\b',
        # Thresholds/limits
        r'\b(?:minimum|maximum|at least|at most|up to|no more than|no less than)\s+(?:of\s+)?\d+(?:\.\d+)?',
        # Eligibility criteria
        r'\b(?:eligible|ineligible|qualifies?|disqualif(?:y|ies|ied)|criteria)\b',
        # Documentation requirements
        r'\b(?:documentation|evidence|supporting|required documents|must provide)\b',
        # Process steps
        r'\b(?:procedure|process|steps?|follow|will be|shall be)\b',
    ]

    # Section header patterns (strongest to weakest)
    HEADER_PATTERNS = [
        # Strong numbered: 1.1.1 Title, 1.1.1.1 Title
        (r'^(\d(?:\.\d+){1,3})\.?\s+(.+)$', 2),
        # Single digit with title: 1. Title, 2. Title
        (r'^(\d)\s*[.:\)]\s*(.+)$', 2),
        # Letter with number: A.1, B.2
        (r'^([A-Z](?:\.\d+)?)\s*[.:\)]\s*(.+)$', 2),
        # ALL CAPS headers (3+ words)
        (r'^([A-Z][A-Z\s]{10,}[A-Z])$', 1),
        # Title Case headers ending with colon
        (r'^([A-Z][a-zA-Z\s]{10,40}:)\s*(.+)$', 2),
    ]

    # Content to skip/ignore
    SKIP_PATTERNS = [
        r'^\s*\d+\s*$',  # Just a number
        r'^page\s+\d+',  # Page numbers
        r'^\s*\d+\s+of\s+\d+\s*$',  # Page X of Y
        r'^(©|copyright|\u00a9)',  # Copyright
        r'^confidential',  # Confidential markers
        r'^(all rights reserved|proprietary)',  # Legal text
        r'^(table of contents|index|appendix)',  # Navigation
    ]

    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.sections: List[GuidelineSection] = []
        self.headers_footers: Set[str] = set()
        self.current_category = "General"
        self.section_counter = 0

    def _is_header_footer(self, text: str, page_num: int) -> bool:
        """Detect if text is a header/footer."""
        text = text.strip()
        if not text:
            return True

        # Common patterns
        patterns = [
            r'^\s*\d+\s*$',  # Just page number
            r'^page\s+\d+',  # "Page 5"
            r'^\d+\s+of\s+\d+',  # "5 of 20"
            r'^(©|copyright)',  # Copyright
            r'^confidential',  # Confidential
            r'^(all rights reserved|proprietary)',  # Legal
            r'^(guidelines?|underwriting)\s+v?\d',  # Document version
        ]

        for pattern in patterns:
            if re.match(pattern, text, re.IGNORECASE):
                return True

        return False

    def _is_section_header(self, text: str) -> Optional[Tuple[str, str, int]]:
        """
        Check if text is a section header.
        Returns (section_id, title, level) or None.
        """
        text = text.strip()
        if len(text) < 5 or len(text) > 200:
            return None

        # Check skip patterns first
        for pattern in self.SKIP_PATTERNS:
            if re.match(pattern, text, re.IGNORECASE):
                return None

        # Try header patterns
        for pattern, group_count in self.HEADER_PATTERNS:
            match = re.match(pattern, text)
            if match:
                if group_count == 2:
                    section_id = match.group(1)
                    title = match.group(2).strip()
                    level = section_id.count('.') if '.' in section_id else 0
                    return section_id, title, level
                else:
                    title = match.group(1).strip()
                    return "", title, 1

        # Check for standalone ALL CAPS (short, standalone lines)
        if text.isupper() and 15 < len(text) < 60:
            words = text.split()
            if len(words) >= 2 and all(w.isalpha() for w in words[:2]):
                return "", text, 1

        return None

    def _is_guideline_content(self, text: str) -> Tuple[bool, str]:
        """
        Determine if text is an actual guideline vs descriptive text.
        Returns (is_guideline, guideline_type).
        """
        text_lower = text.lower()

        # Must be substantial
        if len(text) < 30 or len(text) > 1000:
            return False, ""

        # Check for actual guideline indicators
        for pattern in self.GUIDELINE_INDICATORS:
            if re.search(pattern, text_lower, re.IGNORECASE):
                # Determine type
                if 'must' in text_lower or 'required' in text_lower or 'shall' in text_lower:
                    return True, "requirement"
                elif any(word in text_lower for word in ['may not', 'shall not', 'must not', 'prohibited']):
                    return True, "prohibition"
                elif any(word in text_lower for word in ['minimum', 'maximum', 'at least', 'at most', 'threshold']):
                    return True, "threshold"
                elif any(word in text_lower for word in ['eligible', 'qualifies', 'criteria']):
                    return True, "eligibility"
                else:
                    return True, "procedure"

        # Check for other strong indicators
        strong_indicators = [
            r'\b\d+(?:\.\d+)?%\b',  # Percentages
            r'\$?\d{1,3}(?:,\d{3})+',  # Large numbers (dollar amounts)
            r'\bfico\s+score\b',  # Credit scores
            r'\bltv|cltv|dti\b',  # Ratios
            r'\b(credit score|credit rating)\b',
        ]

        for pattern in strong_indicators:
            if re.search(pattern, text_lower):
                return True, "threshold"

        return False, ""

    def _clean_text(self, text: str) -> str:
        """Clean extracted text."""
        # Normalize whitespace
        text = re.sub(r'\s+', ' ', text)
        # Fix broken hyphenation
        text = re.sub(r'(\w)-\s+(\w)', r'\1\2', text)
        # Remove control characters
        text = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]', '', text)
        return text.strip()

    def _extract_tables(self, page) -> str:
        """Extract table content as formatted text."""
        tables = page.extract_tables()
        if not tables:
            return ""

        result = []
        for table in tables:
            if not table:
                continue
            for row in table:
                if row:
                    # Filter out None values and join
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    if any(cells):
                        result.append(" | ".join(cells))

        return "\n".join(result)

    def extract(self) -> List[GuidelineSection]:
        """Main extraction with CONSERVATIVE approach."""
        print(f"Processing: {self.pdf_path}")

        with pdfplumber.open(self.pdf_path) as pdf:
            total_pages = len(pdf.pages)
            print(f"Total pages: {total_pages}")

            current_section: Optional[GuidelineSection] = None
            section_content_parts: List[str] = []
            section_tables: List[str] = []

            for page_num, page in enumerate(pdf.pages, 1):
                if page_num % 10 == 0:
                    print(f"  Processing page {page_num}/{total_pages}...")

                # Get text
                text = page.extract_text()
                if not text:
                    continue

                # Extract tables
                table_text = self._extract_tables(page)
                if table_text:
                    section_tables.append(f"[Table on page {page_num}]\n{table_text}")

                lines = text.split('\n')
                i = 0
                while i < len(lines):
                    line = lines[i].strip()

                    # Skip headers/footers
                    if self._is_header_footer(line, page_num):
                        i += 1
                        continue

                    # Check for section header
                    header_info = self._is_section_header(line)

                    if header_info:
                        # Save previous section
                        if current_section and section_content_parts:
                            full_content = self._clean_text(' '.join(section_content_parts))
                            if section_tables:
                                full_content += "\n\n" + "\n\n".join(section_tables)
                            current_section.content = full_content

                            # Only keep sections that have actual guideline content
                            if self._has_substantive_content(full_content):
                                self.sections.append(current_section)
                                self.section_counter += 1

                        # Start new section
                        section_id, title, level = header_info
                        if not section_id:
                            section_id = f"SEC{self.section_counter + 1}"

                        # Update category tracking
                        if level == 0 or level == 1:
                            self.current_category = title

                        current_section = GuidelineSection(
                            id=section_id,
                            category=self.current_category,
                            title=title,
                            content="",
                            page_number=page_num,
                            level=level
                        )
                        section_content_parts = []
                        section_tables = []

                    elif current_section is not None:
                        # Accumulate content
                        if line and len(line) > 10:
                            section_content_parts.append(line)

                    i += 1

            # Save final section
            if current_section and section_content_parts:
                full_content = self._clean_text(' '.join(section_content_parts))
                if section_tables:
                    full_content += "\n\n" + "\n\n".join(section_tables)
                current_section.content = full_content

                if self._has_substantive_content(full_content):
                    self.sections.append(current_section)

        print(f"Extracted {len(self.sections)} guideline sections")
        return self.sections

    def _has_substantive_content(self, content: str) -> bool:
        """Check if content has actual guideline value vs just fluff."""
        if not content or len(content) < 50:
            return False

        # Must have some actual content indicators
        words = content.lower().split()

        # Count guideline-related words
        guideline_words = [
            'must', 'shall', 'required', 'minimum', 'maximum', 'eligible',
            'prohibited', 'allowed', 'permitted', 'documentation', 'evidence',
            'criteria', 'requirement', 'limit', 'threshold', 'ratio',
            'score', 'percent', 'percentage', 'amount', 'value'
        ]

        guideline_word_count = sum(1 for w in words if w in guideline_words)

        # Must have at least one guideline word and be substantial
        return guideline_word_count >= 1 and len(content) > 100

    def to_dataframe(self) -> pd.DataFrame:
        """Convert to DataFrame."""
        if not self.sections:
            return pd.DataFrame()

        data = []
        for section in self.sections:
            # Detect guideline type
            is_guideline, g_type = self._is_guideline_content(section.content)

            data.append({
                'id': section.id,
                'category': section.category,
                'title': section.title,
                'content': section.content[:1000] + '...' if len(section.content) > 1000 else section.content,
                'page_number': section.page_number,
                'level': section.level,
                'is_guideline': is_guideline,
                'guideline_type': g_type,
                'word_count': len(section.content.split()),
                'char_count': len(section.content),
            })

        return pd.DataFrame(data)

    def export_to_excel(self, output_path: Optional[str] = None) -> str:
        """Export to formatted Excel."""
        if not self.sections:
            print("No sections to export")
            return ""

        if not output_path:
            base_name = os.path.splitext(os.path.basename(self.pdf_path))[0]
            output_path = f"{base_name}_guidelines.xlsx"

        df = self.to_dataframe()

        wb = Workbook()
        ws = wb.active
        ws.title = "Guidelines"

        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        category_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        headers = ['ID', 'Category', 'Title', 'Content', 'Page', 'Level', 'Type', 'Words']
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='middle')
            cell.border = thin_border

        # Data
        current_category = None
        for idx, row in df.iterrows():
            ws.append([
                row['id'], row['category'], row['title'], row['content'],
                row['page_number'], row['level'], row['guideline_type'], row['word_count']
            ])

            row_num = idx + 2
            if row['category'] != current_category:
                current_category = row['category']
                for col_num in range(1, 9):
                    ws.cell(row=row_num, column=col_num).fill = category_fill

            for col_num in range(1, 9):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 80
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10

        ws.freeze_panes = 'A2'

        # Summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.append(['PDF File:', os.path.basename(self.pdf_path)])
        summary_ws.append(['Total Pages:', df['page_number'].max() if not df.empty else 0])
        summary_ws.append(['Total Sections:', len(self.sections)])
        summary_ws.append(['Categories:', df['category'].nunique() if not df.empty else 0])
        summary_ws.append([])

        # Category breakdown
        summary_ws.append(['Category', 'Count', 'Page Range'])
        for category in sorted(df['category'].unique() if not df.empty else []):
            cat_data = df[df['category'] == category]
            pages = sorted(cat_data['page_number'].unique())
            page_range = f"{min(pages)}-{max(pages)}" if len(pages) > 1 else str(pages[0])
            summary_ws.append([category, len(cat_data), page_range])

        wb.save(output_path)
        print(f"Exported to: {output_path}")
        return output_path

    def export_to_json(self, output_path: Optional[str] = None) -> str:
        """Export to JSON."""
        if not output_path:
            base_name = os.path.splitext(os.path.basename(self.pdf_path))[0]
            output_path = f"{base_name}_guidelines.json"

        data = [asdict(section) for section in self.sections]
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        print(f"Exported to: {output_path}")
        return output_path


def process_pdf(pdf_path: str, output_dir: str = ".") -> Tuple[str, str]:
    """Process a single PDF."""
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    extractor = PDFGuidelineExtractor(pdf_path)
    sections = extractor.extract()

    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    excel_path = os.path.join(output_dir, f"{base_name}_guidelines.xlsx")
    json_path = os.path.join(output_dir, f"{base_name}_guidelines.json")

    extractor.export_to_excel(excel_path)
    extractor.export_to_json(json_path)

    return excel_path, json_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description='Extract guidelines from PDF')
    parser.add_argument('pdf_files', nargs='+', help='PDF file(s) to process')
    parser.add_argument('-o', '--output', default='.', help='Output directory')

    args = parser.parse_args()
    os.makedirs(args.output, exist_ok=True)

    results = []
    for pdf_path in args.pdf_files:
        print(f"\n{'='*60}")
        try:
            excel_path, json_path = process_pdf(pdf_path, args.output)
            results.append({'pdf': pdf_path, 'excel': excel_path, 'json': json_path, 'status': 'success'})
        except Exception as e:
            print(f"ERROR: {e}")
            results.append({'pdf': pdf_path, 'error': str(e), 'status': 'failed'})

    print(f"\n{'='*60}")
    print("SUMMARY")
    successful = sum(1 for r in results if r['status'] == 'success')
    print(f"Successfully processed: {successful}/{len(results)}")


if __name__ == '__main__':
    main()
